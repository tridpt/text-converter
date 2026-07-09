"""Data family converters. Hub representation: a native Python object.

Supported formats: json, yaml, csv, xml.
"""

from __future__ import annotations

import csv
import html as html_lib
import io
import json
import tomllib
import xml.etree.ElementTree as ET
from xml.dom import minidom

import openpyxl
import tomli_w
import yaml
from odf.opendocument import OpenDocumentSpreadsheet
from odf.opendocument import load as odf_load
from odf.table import Table, TableCell, TableRow
from odf.teletype import extractText
from odf.text import P

from .registry import (
    ConversionError,
    FormatSpec,
    reader,
    register_bridge,
    register_format,
    writer,
)

# --- Format catalogue -------------------------------------------------------
register_format(FormatSpec("json", "data", "JSON", ".json", "application/json", False))
register_format(FormatSpec("yaml", "data", "YAML", ".yaml", "application/x-yaml", False))
register_format(FormatSpec("csv", "data", "CSV", ".csv", "text/csv", False))
register_format(FormatSpec("xml", "data", "XML", ".xml", "application/xml", False))
register_format(FormatSpec("toml", "data", "TOML", ".toml", "application/toml", False))
register_format(
    FormatSpec(
        "xlsx",
        "data",
        "Excel (XLSX)",
        ".xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        True,
    )
)
register_format(
    FormatSpec(
        "ods",
        "data",
        "OpenDocument Sheet (ODS)",
        ".ods",
        "application/vnd.oasis.opendocument.spreadsheet",
        True,
    )
)


def _decode(data: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


# --- Readers (bytes -> Python object) ---------------------------------------
@reader("json")
def read_json(data: bytes) -> object:
    try:
        return json.loads(_decode(data))
    except json.JSONDecodeError as exc:
        raise ConversionError(f"Invalid JSON: {exc}") from exc


@reader("yaml")
def read_yaml(data: bytes) -> object:
    try:
        return yaml.safe_load(_decode(data))
    except yaml.YAMLError as exc:
        raise ConversionError(f"Invalid YAML: {exc}") from exc


@reader("csv")
def read_csv(data: bytes) -> object:
    text = _decode(data)
    rows = list(csv.DictReader(io.StringIO(text)))
    return rows


@reader("xml")
def read_xml(data: bytes) -> object:
    try:
        root = ET.fromstring(_decode(data))
    except ET.ParseError as exc:
        raise ConversionError(f"Invalid XML: {exc}") from exc
    return {root.tag: _elem_to_obj(root)}


@reader("toml")
def read_toml(data: bytes) -> object:
    try:
        return tomllib.loads(_decode(data))
    except tomllib.TOMLDecodeError as exc:
        raise ConversionError(f"Invalid TOML: {exc}") from exc


@writer("toml")
def write_toml(obj: object, options=None) -> bytes:
    if not isinstance(obj, dict):
        obj = {"data": obj}
    try:
        return tomli_w.dumps(obj).encode("utf-8")
    except (TypeError, ValueError) as exc:
        raise ConversionError(
            f"Cannot represent this data as TOML: {exc}. "
            "TOML requires a table (mapping) structure without null values."
        ) from exc


def _elem_to_obj(elem: ET.Element) -> object:
    children = list(elem)
    if not children:
        text = (elem.text or "").strip()
        return text
    result: dict = {}
    for child in children:
        value = _elem_to_obj(child)
        if child.tag in result:
            if not isinstance(result[child.tag], list):
                result[child.tag] = [result[child.tag]]
            result[child.tag].append(value)
        else:
            result[child.tag] = value
    return result


# --- Writers (Python object -> bytes) ---------------------------------------
@writer("json")
def write_json(obj: object, options=None) -> bytes:
    return json.dumps(obj, indent=2, ensure_ascii=False).encode("utf-8")


@writer("yaml")
def write_yaml(obj: object, options=None) -> bytes:
    return yaml.safe_dump(obj, allow_unicode=True, sort_keys=False).encode("utf-8")


@writer("csv")
def write_csv(obj: object, options=None) -> bytes:
    rows = _as_rows(obj)
    if not rows:
        return b""
    fieldnames: list[str] = []
    for row in rows:
        for key in row:
            if key not in fieldnames:
                fieldnames.append(key)
    buf = io.StringIO()
    dict_writer = csv.DictWriter(buf, fieldnames=fieldnames)
    dict_writer.writeheader()
    for row in rows:
        dict_writer.writerow({k: _scalar(row.get(k, "")) for k in fieldnames})
    return buf.getvalue().encode("utf-8")


def _as_rows(obj: object) -> list[dict]:
    """Coerce an arbitrary object into a list of flat dict rows for CSV."""
    if isinstance(obj, list):
        return [r if isinstance(r, dict) else {"value": r} for r in obj]
    if isinstance(obj, dict):
        # A dict whose single value is a list of records -> use that list.
        if len(obj) == 1:
            (only_value,) = obj.values()
            if isinstance(only_value, list):
                return _as_rows(only_value)
        return [obj]
    return [{"value": obj}]


def _scalar(value: object) -> str:
    if isinstance(value, dict | list):
        return json.dumps(value, ensure_ascii=False)
    return "" if value is None else str(value)


@writer("xml")
def write_xml(obj: object, options=None) -> bytes:
    if isinstance(obj, dict) and len(obj) == 1:
        (root_tag,) = obj.keys()
        (root_value,) = obj.values()
        root = ET.Element(_safe_tag(root_tag))
        _build_xml(root, root_value)
    else:
        root = ET.Element("root")
        _build_xml(root, obj)
    rough = ET.tostring(root, encoding="utf-8")
    pretty = minidom.parseString(rough).toprettyxml(indent="  ", encoding="utf-8")
    return pretty


def _build_xml(parent: ET.Element, value: object) -> None:
    if isinstance(value, dict):
        for key, val in value.items():
            tag = _safe_tag(str(key))
            if isinstance(val, list):
                for item in val:
                    child = ET.SubElement(parent, tag)
                    _build_xml(child, item)
            else:
                child = ET.SubElement(parent, tag)
                _build_xml(child, val)
    elif isinstance(value, list):
        for item in value:
            child = ET.SubElement(parent, "item")
            _build_xml(child, item)
    else:
        parent.text = "" if value is None else str(value)


def _safe_tag(tag: str) -> str:
    tag = tag.strip().replace(" ", "_")
    if not tag or not (tag[0].isalpha() or tag[0] == "_"):
        tag = "_" + tag
    return "".join(c for c in tag if c.isalnum() or c in "_-.")


# --- Spreadsheets: XLSX (openpyxl) ------------------------------------------
@reader("xlsx")
def read_xlsx(data: bytes) -> object:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ConversionError(f"Invalid XLSX file: {exc}") from exc
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return _grid_to_rows(rows)


@writer("xlsx")
def write_xlsx(obj: object, options=None) -> bytes:
    rows = _as_rows(obj)
    wb = openpyxl.Workbook()
    ws = wb.active
    columns = _columns_of(rows)
    if columns:
        ws.append(columns)
        for row in rows:
            ws.append([_scalar(row.get(c, "")) for c in columns])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- Spreadsheets: ODS (odfpy) ----------------------------------------------
@reader("ods")
def read_ods(data: bytes) -> object:
    try:
        doc = odf_load(io.BytesIO(data))
    except Exception as exc:  # noqa: BLE001
        raise ConversionError(f"Invalid ODS file: {exc}") from exc
    tables = doc.getElementsByType(Table)
    if not tables:
        return []
    grid = []
    for tr in tables[0].getElementsByType(TableRow):
        values = []
        for cell in tr.getElementsByType(TableCell):
            repeat = min(int(cell.getAttribute("numbercolumnsrepeated") or 1), 1024)
            text = "".join(extractText(p) for p in cell.getElementsByType(P))
            values.extend([text] * repeat)
        grid.append(values)
    return _grid_to_rows(grid)


@writer("ods")
def write_ods(obj: object, options=None) -> bytes:
    rows = _as_rows(obj)
    doc = OpenDocumentSpreadsheet()
    table = Table(name="Sheet1")
    columns = _columns_of(rows)

    def _add(values):
        tr = TableRow()
        for value in values:
            cell = TableCell()
            cell.addElement(P(text=_scalar(value)))
            tr.addElement(cell)
        table.addElement(tr)

    if columns:
        _add(columns)
        for row in rows:
            _add([row.get(c, "") for c in columns])
    doc.spreadsheet.addElement(table)
    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


def _columns_of(rows: list[dict]) -> list[str]:
    columns: list[str] = []
    for row in rows:
        for key in row:
            if key not in columns:
                columns.append(key)
    return columns


def _grid_to_rows(grid) -> list[dict]:
    """Turn a 2-D grid (list of row lists/tuples) into a list of dict records."""
    grid = [list(r) for r in grid]
    while grid and all((v is None or v == "") for v in grid[-1]):
        grid.pop()
    if not grid:
        return []
    header = [("" if h is None else str(h)) for h in grid[0]]
    while header and header[-1] == "":
        header.pop()
    ncols = len(header)
    out = []
    for row in grid[1:]:
        record = {}
        for i in range(ncols):
            value = row[i] if i < len(row) else ""
            record[header[i]] = "" if value is None else value
        out.append(record)
    return out


# --- Bridge: data family -> document family ---------------------------------
# Renders a Python object into an HTML fragment (the document family hub) so
# that data formats can be exported to txt/md/html/docx/pdf/... as well.
def _obj_to_html(obj: object) -> str:
    return f"<h1>Data</h1>\n{_render(obj)}"


def _render(value: object) -> str:
    if isinstance(value, list):
        if value and all(isinstance(item, dict) for item in value):
            return _render_table(value)
        items = "".join(f"<li>{_render(item)}</li>" for item in value)
        return f"<ul>{items}</ul>"
    if isinstance(value, dict):
        rows = ""
        for key, val in value.items():
            rows += (
                f"<tr><th style='text-align:left'>{html_lib.escape(str(key))}</th>"
                f"<td>{_render(val)}</td></tr>"
            )
        return f"<table border='1' cellpadding='4'>{rows}</table>"
    if value is None:
        return "<em>null</em>"
    return html_lib.escape(str(value))


def _render_table(rows: list[dict]) -> str:
    columns: list[str] = []
    for row in rows:
        for key in row:
            if key not in columns:
                columns.append(key)
    head = "".join(f"<th>{html_lib.escape(str(c))}</th>" for c in columns)
    body = ""
    for row in rows:
        cells = "".join(f"<td>{_render(row.get(c, ''))}</td>" for c in columns)
        body += f"<tr>{cells}</tr>"
    return (
        "<table border='1' cellpadding='4'>"
        f"<thead><tr>{head}</tr></thead><tbody>{body}</tbody></table>"
    )


register_bridge("data", "document", _obj_to_html)
